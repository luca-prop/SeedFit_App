#!/usr/bin/env python3
"""Patch selected Golden Sample rows (stage/name/notes) via Drive xlsx."""

from __future__ import annotations

import argparse
import json
import shutil
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
from apply_jaegebal_district_status import (  # noqa: E402
    GOLDEN_SHEET_ID,
    run,
    upload_xlsx_replace,
    zone_key,
)
from normalize_golden_samples import derive_coverage  # noqa: E402
from reorder_golden_by_coverage import reorder_workbook  # noqa: E402


def apply_patches(src: Path, out: Path, patches: list[dict[str, Any]]) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(src)
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]
    hm = {str(h): i + 1 for i, h in enumerate(headers) if h}
    by_patch = {p["naturalKey"]: p for p in patches}
    found: list[str] = []
    renamed: list[str] = []

    for r in range(2, ws.max_row + 1):
        key = zone_key(
            ws.cell(r, hm["행정구"]).value,
            ws.cell(r, hm["행정동"]).value,
            ws.cell(r, hm["구역명"]).value,
        )
        patch = by_patch.get(key)
        if not patch:
            continue
        found.append(key)
        if patch.get("renameZone"):
            ws.cell(r, hm["구역명"], patch["renameZone"])
            renamed.append(f"{key} → {patch['renameZone']}")
        if patch.get("현재 단계") is not None:
            ws.cell(r, hm["현재 단계"], patch["현재 단계"])
        if patch.get("coverage") is not None:
            ws.cell(r, hm["coverage"], patch["coverage"])
        elif patch.get("현재 단계"):
            ws.cell(r, hm["coverage"], derive_coverage(patch["현재 단계"]))
        if patch.get("특징/호재") is not None and "특징/호재" in hm:
            ws.cell(r, hm["특징/호재"], patch["특징/호재"])
        if patch.get("jaegebal_url") is not None and "jaegebal_url" in hm:
            ws.cell(r, hm["jaegebal_url"], patch["jaegebal_url"])
        for col in (
            "최소 실투자금(억)",
            "최대 실투자금(억)",
            "최소 프리미엄",
            "최대 프리미엄",
        ):
            if patch.get(col) is not None and col in hm:
                ws.cell(r, hm[col], patch[col])

    missing = [k for k in by_patch if k not in found]
    wb.save(out)
    return {"patched": found, "renamed": renamed, "missing": missing, "output": str(out)}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--patches", type=Path, required=True)
    parser.add_argument("--sheet-id", default=GOLDEN_SHEET_ID)
    parser.add_argument("--remote", default="gdrive")
    parser.add_argument("--workdir", type=Path, default=Path("data/tmp_patch_golden_rows"))
    parser.add_argument("--write", action="store_true")
    parser.add_argument("--no-reorder", action="store_true")
    args = parser.parse_args()

    patches = json.loads(args.patches.read_text(encoding="utf-8"))
    if not isinstance(patches, list):
        raise SystemExit("patches JSON must be a list")

    workdir = args.workdir
    if workdir.exists():
        shutil.rmtree(workdir)
    workdir.mkdir(parents=True)
    run(
        [
            "rclone",
            "backend",
            "copyid",
            f"{args.remote}:",
            args.sheet_id,
            f"{workdir}/",
            "--drive-export-formats",
            "xlsx",
        ]
    )
    files = sorted(workdir.glob("*.xlsx"))
    if not files:
        raise SystemExit("no xlsx exported")
    patched = workdir / "patched.xlsx"
    stats = apply_patches(files[0], patched, patches)
    final = patched
    if not args.no_reorder:
        reordered = workdir / "patched_reordered.xlsx"
        stats["reorder"] = reorder_workbook(patched, reordered)
        final = reordered
    print(json.dumps({**stats, "write": bool(args.write)}, ensure_ascii=False, indent=2))
    if stats.get("missing"):
        raise SystemExit(f"missing keys: {stats['missing']}")
    if args.write:
        upload_xlsx_replace(args.sheet_id, final, remote=args.remote)
        print(f"https://docs.google.com/spreadsheets/d/{args.sheet_id}/edit")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
