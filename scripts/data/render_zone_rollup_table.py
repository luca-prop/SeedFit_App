#!/usr/bin/env python3
"""Render the 구역 통합본 (1차 승인 단위) as a double-clickable HTML table + xlsx.

The listing-review board is a working UI. This file is the *handoff artifact*:
초투 / 매매가(호가) / P in one table that opens in Edge/Chrome by double-click
(ASCII filename + complete html/head/body). Also writes a dated archive xlsx
so next month's update can diff 초투·매매가.
"""

from __future__ import annotations

import argparse
import csv
import html
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[2]
ARCHIVE_ROOT = ROOT / "data" / "archive" / "golden"

# HTML·xlsx 동일 열. C=구역 다음이 현재 단계. 초투·매매는 min/max 숫자.
TABLE_COLUMNS: list[tuple[str, str, str]] = [
    ("행정구", "구", "text"),
    ("행정동", "동", "text"),
    ("구역명", "구역", "text"),
    ("현재 단계", "현재 단계", "text"),
    ("표본건수", "표본", "int"),
    ("예상초투_min(억)", "초투min", "num"),
    ("예상초투_max(억)", "초투max", "num"),
    ("매매가_min(억)", "매매min", "num"),
    ("매매가_max(억)", "매매max", "num"),
    ("P_min(억)", "Pmin", "num"),
    ("P_max(억)", "Pmax", "num"),
    ("기존_최소실투자금(억)", "기존실투min", "num"),
    ("기존_최대실투자금(억)", "기존실투max", "num"),
    ("기존_매매가", "기존매매가", "text"),
    ("zoneNaturalKey", "zoneNaturalKey", "text"),
]
COLUMNS = [(key, label) for key, label, _kind in TABLE_COLUMNS]
XLSX_COLUMNS = TABLE_COLUMNS


def _cell(row: dict[str, str], key: str) -> str:
    return (row.get(key) or "").strip() or "—"


def _eok_number(value: str | None) -> float | None:
    """CSV 억 칸 → Excel 숫자. 범위 문자열·공란은 None."""
    text = (value or "").replace("억", "").replace(",", "").strip()
    if not text or text in {"—", "-", "-"}:
        return None
    if "~" in text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _int_or_none(value: str | None) -> int | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def load_stage_index(golden_csv: Path | None) -> dict[str, str]:
    if golden_csv is None or not golden_csv.exists():
        return {}
    out: dict[str, str] = {}
    for row in load_csv(golden_csv):
        stage = (row.get("현재 단계") or "").strip()
        name = (row.get("구역명") or "").strip()
        key = (row.get("zoneNaturalKey") or "").strip()
        if not key:
            key = "|".join(
                [
                    (row.get("행정구") or "").strip(),
                    (row.get("행정동") or "").strip(),
                    name,
                ]
            )
        if key and stage:
            out[key] = stage
        if name and stage:
            out[name] = stage
    return out


def prepare_rollup_rows(
    rows: list[dict[str, str]],
    golden_csv: Path | None = None,
) -> list[dict[str, str]]:
    stages = load_stage_index(golden_csv)
    prepared: list[dict[str, str]] = []
    for row in rows:
        item = dict(row)
        key = (item.get("zoneNaturalKey") or "").strip()
        name = (item.get("구역명") or "").strip()
        if not (item.get("현재 단계") or "").strip():
            item["현재 단계"] = stages.get(key) or stages.get(name) or ""
        prepared.append(item)

    def sort_key(row: dict[str, str]) -> tuple[int, float]:
        value = _eok_number(row.get("예상초투_min(억)"))
        if value is None:
            return (1, 0.0)
        return (0, value)

    return sorted(prepared, key=sort_key)


def _html_value(row: dict[str, str], key: str, kind: str) -> str:
    if kind == "num":
        number = _eok_number(row.get(key))
        if number is None:
            return "—"
        if number == int(number):
            return str(int(number))
        return f"{number:g}"
    if kind == "int":
        number = _int_or_none(row.get(key))
        return "—" if number is None else str(number)
    return html.escape(_cell(row, key))


def render_html(
    rows: list[dict[str, str]],
    *,
    stamp: str,
    source: str,
    golden_csv: Path | None = None,
) -> str:
    rows = prepare_rollup_rows(rows, golden_csv)
    body_rows = []
    for row in rows:
        tds = []
        for key, _label, kind in TABLE_COLUMNS:
            value = _html_value(row, key, kind)
            if key == "구역명":
                value = f"<strong>{value}</strong>"
            css = ' class="num"' if kind in {"num", "int"} else ""
            tds.append(f"<td{css}>{value}</td>")
        body_rows.append("<tr>" + "".join(tds) + "</tr>")
    header = "".join(f"<th>{html.escape(label)}</th>" for _key, label, _kind in TABLE_COLUMNS)
    iso = f"20{stamp[:2]}-{stamp[2:4]}-{stamp[4:6]}" if len(stamp) == 6 else stamp
    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta http-equiv="X-UA-Compatible" content="IE=edge">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>구역 통합본 (1차 승인 단위) {iso}</title>
  <style>
    body {{ font: 15px/1.45 "Malgun Gothic", "Apple SD Gothic Neo", sans-serif; margin: 24px; color: #111; background: #fff; }}
    h1 {{ font-size: 22px; margin: 0 0 8px; }}
    .note {{ color: #444; max-width: 80rem; margin: 0 0 16px; }}
    table {{ border-collapse: collapse; width: 100%; }}
    th, td {{ border: 1px solid #ccc; padding: 7px 8px; text-align: left; white-space: nowrap; }}
    th {{ background: #f3efe4; position: sticky; top: 0; }}
    td.num {{ text-align: right; font-variant-numeric: tabular-nums; }}
    tr:nth-child(even) {{ background: #fafafa; }}
  </style>
</head>
<body>
  <h1>구역 통합본 (1차 승인 단위) — {html.escape(iso)}</h1>
  <p class="note">
    매물 검수보드 승인분 기준. 행은 <b>초투min</b> 오름차순.
    초투·매매는 min/max 숫자. 탐색기에서 더블클릭하면 브라우저로 열립니다.
    출처: {html.escape(source)}. P는 매물 설명의 프리미엄/프미/P/피 힌트이며, 힌트가 없으면 공란입니다.
  </p>
  <table>
    <thead><tr>{header}</tr></thead>
    <tbody>
      {"".join(body_rows)}
    </tbody>
  </table>
</body>
</html>
"""


def write_xlsx(
    rows: list[dict[str, str]],
    path: Path,
    golden_csv: Path | None = None,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "구역통합본"
    fill = PatternFill("solid", fgColor="F3EFE4")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    rows = prepare_rollup_rows(rows, golden_csv)
    ws.append([label for _key, label, _kind in XLSX_COLUMNS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        values: list[object] = []
        for key, _label, kind in XLSX_COLUMNS:
            raw = row.get(key)
            if kind == "num":
                values.append(_eok_number(raw))
            elif kind == "int":
                values.append(_int_or_none(raw))
            else:
                text = (raw or "").strip()
                values.append(text if text else "—")
        ws.append(values)
    for excel_row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in excel_row:
            cell.border = thin
    num_cols = [i for i, (_k, _l, kind) in enumerate(XLSX_COLUMNS, start=1) if kind == "num"]
    for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for col_idx in num_cols:
            cell = excel_row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.##"
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    widths = [10, 12, 28, 16, 8, 10, 10, 10, 10, 10, 10, 12, 12, 14, 36]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def fmt_sale_cell(lo: str, hi: str) -> str:
    """Golden G열 표기: 자양7은 `17~18`. 단일이면 숫자만."""
    lo, hi = (lo or "").strip(), (hi or "").strip()
    if not lo and not hi:
        return ""
    if not hi or lo == hi:
        return lo or hi
    if not lo:
        return hi
    return f"{lo}~{hi}"


def patches_from_rollup(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    patches: list[dict[str, Any]] = []
    for row in rows:
        key = (row.get("zoneNaturalKey") or "").strip()
        n = int(row.get("표본건수") or 0)
        if not key or n <= 0 or not (row.get("예상초투_min(억)") or "").strip():
            continue
        patch: dict[str, Any] = {
            "naturalKey": key,
            "최소 실투자금(억)": row["예상초투_min(억)"].strip(),
            "최대 실투자금(억)": row["예상초투_max(억)"].strip(),
            "매매가": fmt_sale_cell(row.get("매매가_min(억)") or "", row.get("매매가_max(억)") or ""),
        }
        if (row.get("P_min(억)") or "").strip():
            patch["최소 프리미엄"] = row["P_min(억)"].strip()
        if (row.get("P_max(억)") or "").strip():
            patch["최대 프리미엄"] = row["P_max(억)"].strip()
        patches.append(patch)
    return patches


def copy_rollup_artifacts(
    stamp: str,
    *,
    html_path: Path,
    xlsx_path: Path,
    golden_xlsx: Path | None = None,
) -> None:
    """루트(최신) + YYMMDD 폴더에 html/xlsx/golden 스냅샷을 둔다."""
    ARCHIVE_ROOT.mkdir(parents=True, exist_ok=True)
    dated = ARCHIVE_ROOT / stamp
    dated.mkdir(parents=True, exist_ok=True)
    copies = [
        (html_path, ARCHIVE_ROOT / "zone_rollup.html"),
        (html_path, dated / f"zone_rollup_{stamp}.html"),
        (xlsx_path, ARCHIVE_ROOT / "zone_rollup.xlsx"),
        (xlsx_path, dated / f"zone_rollup_{stamp}.xlsx"),
    ]
    if golden_xlsx is not None and golden_xlsx.exists():
        copies.extend(
            [
                (golden_xlsx, ARCHIVE_ROOT / "golden_samples.xlsx"),
                (golden_xlsx, dated / f"golden_samples_{stamp}.xlsx"),
            ]
        )
    for src, dest in copies:
        try:
            dest.write_bytes(src.read_bytes())
            print("archive", dest)
        except OSError as exc:
            print(f"archive skip {dest} ({exc})")


def snapshot_golden_xlsx(golden_csv: Path, dest: Path) -> None:
    from openpyxl import Workbook

    rows = load_csv(golden_csv)
    if not rows:
        raise SystemExit(f"empty golden csv: {golden_csv}")
    wb = Workbook()
    ws = wb.active
    ws.title = "golden"
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, required=True, help="구역_통합본_YYMMDD.csv")
    parser.add_argument("--stamp", default=date.today().strftime("%y%m%d"))
    parser.add_argument("--golden-csv", type=Path, help="optional: snapshot this Golden CSV to archive xlsx")
    parser.add_argument("--desktop", action="store_true", help="also copy HTML to the user Desktop")
    args = parser.parse_args()

    rows = load_csv(args.input)
    stamp = args.stamp
    reports = ROOT / "data" / "reports"
    archive = ROOT / "data" / "archive" / "golden" / stamp
    archive.mkdir(parents=True, exist_ok=True)

    html_name = f"zone_rollup_{stamp}.html"
    html_body = render_html(
        rows, stamp=stamp, source=args.input.name, golden_csv=args.golden_csv
    )
    reports.mkdir(parents=True, exist_ok=True)
    html_reports = reports / html_name
    html_reports.write_text(html_body, encoding="utf-8")
    (reports / f"구역_통합본_{stamp}.html").write_text(html_body, encoding="utf-8")
    print("html", html_reports)

    xlsx_reports = reports / f"zone_rollup_{stamp}.xlsx"
    write_xlsx(rows, xlsx_reports, golden_csv=args.golden_csv)
    print("xlsx", xlsx_reports)

    patch_path = reports / f"golden_invest_p_patches_{stamp}.json"
    import json

    patches = patches_from_rollup(rows)
    patch_path.write_text(json.dumps(patches, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"patches {len(patches)} → {patch_path}")

    golden_xlsx = None
    if args.golden_csv and args.golden_csv.exists():
        golden_xlsx = reports / f"golden_samples_{stamp}.xlsx"
        snapshot_golden_xlsx(args.golden_csv, golden_xlsx)
        print("golden snapshot", golden_xlsx)

    copy_rollup_artifacts(
        stamp,
        html_path=html_reports,
        xlsx_path=xlsx_reports,
        golden_xlsx=golden_xlsx,
    )

    if args.desktop:
        desktop = Path.home() / "Desktop" / html_name
        desktop.write_text(html_body, encoding="utf-8")
        print("desktop", desktop)


if __name__ == "__main__":
    main()
