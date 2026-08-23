#!/usr/bin/env python3
"""Diff two dated Golden archives on 초투 / 매매가 / P."""

from __future__ import annotations

import argparse
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]


def _num(value: object) -> float | None:
    text = str(value or "").replace("억", "").replace(",", "").strip()
    if not text or text == "—":
        return None
    if "~" in text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def load_rollup(stamp: str) -> dict[str, dict]:
    from openpyxl import load_workbook

    path = ROOT / "data" / "archive" / "golden" / stamp / f"zone_rollup_{stamp}.xlsx"
    if not path.exists():
        raise SystemExit(f"missing {path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value or "") for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    key_col = idx.get("zoneNaturalKey")
    if key_col is None:
        raise SystemExit("zoneNaturalKey column missing")
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = str(row[key_col] or "")
        if not key:
            continue
        out[key] = {h: row[i] for i, h in enumerate(headers)}
    return out


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--before", required=True, help="YYMMDD")
    parser.add_argument("--after", required=True, help="YYMMDD")
    args = parser.parse_args()
    before = load_rollup(args.before)
    after = load_rollup(args.after)
    fields = [
        ("초투min", "초투min"),
        ("초투max", "초투max"),
        ("매매min", "매매min"),
        ("매매max", "매매max"),
        ("Pmin", "Pmin"),
        ("Pmax", "Pmax"),
    ]
    print(f"key\t" + "\t".join(f"{label}_Δ" for _k, label in fields))
    for key in sorted(set(before) | set(after)):
        b, a = before.get(key, {}), after.get(key, {})
        deltas = []
        changed = False
        for col, _label in fields:
            bv, av = _num(b.get(col)), _num(a.get(col))
            if bv is None and av is None:
                deltas.append("")
            elif bv is None or av is None:
                deltas.append(f"{bv}->{av}")
                changed = True
            else:
                d = round(av - bv, 4)
                deltas.append(str(d) if d else "0")
                if d:
                    changed = True
        if changed:
            print(key + "\t" + "\t".join(deltas))


if __name__ == "__main__":
    main()
