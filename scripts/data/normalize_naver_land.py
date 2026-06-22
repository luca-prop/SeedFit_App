#!/usr/bin/env python3
"""Normalize Naver Land XLSX files into MVP reference apartment payload JSON."""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

EOK_KRW = Decimal("100000000")
EXPECTED_HEADERS = ["단지명", "구분", "거래", "전용(m2)", "층", "가격(억)", "특징"]
REQUIRED_SOURCE_FILES = {"Naver_Land_0503_1129.xlsx", "Naver_Land_0503_1132.xlsx"}
FLOOR_TIER_BY_TEXT = {
    "상": "HIGH",
    "고": "HIGH",
    "중": "MID",
    "하": "LOW",
    "저": "LOW",
}


@dataclass
class WarningItem:
    sourceFile: str
    row: int
    code: str
    message: str
    apartmentName: str | None = None


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = re.sub(r"[ \t]+", " ", str(value).strip())
    return cleaned or None


def normalize_apartment_name(value: Any) -> str | None:
    text = clean_text(value)
    if text is None:
        return None
    return text.replace(" ", "")


def source_captured_at_from_filename(path: Path, year: int) -> str:
    match = re.search(r"_(\d{2})(\d{2})_(\d{2})(\d{2})", path.stem)
    if not match:
        raise ValueError(f"Cannot infer source capture time from filename: {path.name}")
    month, day, hour, minute = match.groups()
    return f"{year}-{month}-{day}T{hour}:{minute}:00+09:00"


def parse_eok_to_krw(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, int | float | Decimal):
        return int(Decimal(str(value)) * EOK_KRW)
    normalized = str(value).replace(",", "").replace("억원", "").replace("억", "").strip()
    if not normalized:
        return None
    try:
        return int(Decimal(normalized) * EOK_KRW)
    except InvalidOperation as exc:
        raise ValueError(f"Invalid eok money value: {value}") from exc


def parse_area_m2(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except InvalidOperation as exc:
        raise ValueError(f"Invalid area value: {value}") from exc


def decimal_to_json(value: Decimal | None) -> float | None:
    if value is None:
        return None
    return float(value)


def parse_floor(value: Any) -> dict[str, Any]:
    text = clean_text(value)
    if text is None:
        return {"rawFloor": None, "floor": None, "totalFloor": None, "floorBand": None, "floorTier": None}
    parts = text.split("/", maxsplit=1)
    floor_part = parts[0]
    total_floor = int(parts[1]) if len(parts) == 2 and parts[1].isdigit() else None
    if floor_part.isdigit():
        floor = int(floor_part)
        floor_band = None
    else:
        floor = None
        floor_band = floor_part
    floor_tier = FLOOR_TIER_BY_TEXT.get(floor_band) if floor_band else None
    return {
        "rawFloor": text,
        "floor": floor,
        "totalFloor": total_floor,
        "floorBand": floor_band,
        "floorTier": floor_tier,
    }


def validate_source_files(paths: list[Path]) -> None:
    actual_files = {path.name for path in paths}
    if actual_files != REQUIRED_SOURCE_FILES:
        raise ValueError(
            "Naver Land normalization requires exactly these two files: "
            f"{sorted(REQUIRED_SOURCE_FILES)}. Got: {sorted(actual_files)}"
        )


def normalize(paths: list[Path], output_path: Path, year: int) -> dict[str, Any]:
    validate_source_files(paths)
    listings: list[dict[str, Any]] = []
    warnings: list[WarningItem] = []
    without_current_listing = 0
    excluded_low_floor = 0

    for path in paths:
        source_captured_at = source_captured_at_from_filename(path, year)
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = [clean_text(value) for value in next(rows)]

        if headers != EXPECTED_HEADERS:
            warnings.append(
                WarningItem(
                    sourceFile=path.name,
                    row=1,
                    code="unexpected_headers",
                    message=f"Expected {EXPECTED_HEADERS}, got {headers}",
                )
            )
            continue

        header_index = {header: index for index, header in enumerate(headers)}
        for row_index, row in enumerate(rows, start=2):
            apartment_name = normalize_apartment_name(row[header_index["단지명"]])
            listing_type = clean_text(row[header_index["구분"]])
            deal_type = clean_text(row[header_index["거래"]])
            raw_notes = clean_text(row[header_index["특징"]])

            if apartment_name is None:
                warnings.append(WarningItem(path.name, row_index, "missing_apartment_name", "apartment name is missing"))
                continue
            if listing_type == "매물없음":
                without_current_listing += 1
                continue
            if deal_type != "매매":
                warnings.append(WarningItem(path.name, row_index, "non_sale_listing", f"unsupported deal type: {deal_type}", apartment_name))
                continue

            try:
                area_m2 = parse_area_m2(row[header_index["전용(m2)"]])
                price_krw = parse_eok_to_krw(row[header_index["가격(억)"]])
            except ValueError as exc:
                warnings.append(WarningItem(path.name, row_index, "parse_error", str(exc), apartment_name))
                continue

            if area_m2 is None:
                warnings.append(WarningItem(path.name, row_index, "missing_area_m2", "exclusive area is missing", apartment_name))
            if price_krw is None:
                warnings.append(WarningItem(path.name, row_index, "missing_price", "price is missing", apartment_name))
                continue

            floor = parse_floor(row[header_index["층"]])
            is_presale = "분양권" in f"{apartment_name} {raw_notes or ''}"
            is_reference_price_candidate = listing_type != "저층"
            if not is_reference_price_candidate:
                excluded_low_floor += 1

            listings.append(
                {
                    "naturalKey": f"{apartment_name}|{decimal_to_json(area_m2)}|{is_presale}|{row_index}|{path.name}",
                    "apartmentName": apartment_name,
                    "areaM2": decimal_to_json(area_m2),
                    "listingType": listing_type,
                    "dealType": deal_type,
                    "priceKrw": price_krw,
                    "isReferencePriceCandidate": is_reference_price_candidate,
                    "referencePriceExclusionReason": None if is_reference_price_candidate else "low_floor",
                    "isPresale": is_presale,
                    "rawFloor": floor["rawFloor"],
                    "floor": floor["floor"],
                    "totalFloor": floor["totalFloor"],
                    "floorBand": floor["floorBand"],
                    "floorTier": floor["floorTier"],
                    "notes": raw_notes,
                    "sourceFile": path.name,
                    "sourceCapturedAt": source_captured_at,
                    "sourceRow": row_index,
                }
            )

    reference_apartments: list[dict[str, Any]] = []
    low_floor_fallback_apartments = []
    apartment_names = sorted({listing["apartmentName"] for listing in listings})
    for apartment_name in apartment_names:
        apartment_listings = [listing for listing in listings if listing["apartmentName"] == apartment_name]
        presale_values = sorted({listing["isPresale"] for listing in apartment_listings})
        for is_presale in presale_values:
            same_apartment_listings = [listing for listing in apartment_listings if listing["isPresale"] == is_presale]
            preferred_listings = [listing for listing in same_apartment_listings if listing["isReferencePriceCandidate"]]
            is_low_floor_fallback = not preferred_listings
            items = preferred_listings or same_apartment_listings
            if is_low_floor_fallback:
                low_floor_fallback_apartments.append(
                    {
                        "apartmentName": apartment_name,
                        "listingCount": len(same_apartment_listings),
                        "reason": "only_low_floor_listings",
                    }
                )

            representative = sorted(items, key=lambda item: (item["priceKrw"], item["sourceCapturedAt"], item["sourceRow"]))[0]
            prices = sorted(item["priceKrw"] for item in items)
            reference_apartments.append(
                {
                    "naturalKey": f"{representative['apartmentName']}|{representative['areaM2']}|{representative['isPresale']}",
                    "apartmentName": representative["apartmentName"],
                    "district": None,
                    "dong": None,
                    "areaM2": representative["areaM2"],
                    "currentPriceKrw": representative["priceKrw"],
                    "priceMinKrw": min(prices),
                    "priceMaxKrw": max(prices),
                    "priceSelectionPolicy": "lowest_low_floor_fallback" if is_low_floor_fallback else "lowest_non_low_floor_listing",
                    "isLowFloorFallback": is_low_floor_fallback,
                    "listingCount": len(items),
                    "allListingCount": sum(
                        1
                        for listing in listings
                        if listing["apartmentName"] == representative["apartmentName"]
                        and listing["isPresale"] == representative["isPresale"]
                    ),
                    "selectedSourceRow": representative["sourceRow"],
                    "selectedListingType": representative["listingType"],
                    "selectedRawFloor": representative["rawFloor"],
                    "isPresale": representative["isPresale"],
                    "sourceFiles": sorted({item["sourceFile"] for item in items}),
                    "sourceCapturedAt": max(item["sourceCapturedAt"] for item in items),
                }
            )

    warnings_payload = [warning.__dict__ for warning in warnings]
    payload = {
        "source": {
            "files": [path.as_posix() for path in paths],
            "sourceCapturedAt": sorted({source_captured_at_from_filename(path, year) for path in paths}),
            "sourceYear": year,
        },
        "referenceApartments": reference_apartments,
        "referenceApartmentListingEvidence": listings,
        "lowFloorFallbackApartments": low_floor_fallback_apartments,
        "summary": {
            "sourceFileCount": len(paths),
            "listingCount": len(listings),
            "referencePriceCandidateCount": sum(1 for listing in listings if listing["isReferencePriceCandidate"]),
            "excludedLowFloorListingCount": excluded_low_floor,
            "referenceApartmentCount": len(reference_apartments),
            "lowFloorFallbackCount": len(low_floor_fallback_apartments),
            "withoutCurrentListing": without_current_listing,
            "warningCount": len(warnings_payload),
            "warnings": warnings_payload,
        },
    }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize Naver Land XLSX files.")
    parser.add_argument(
        "--input",
        nargs="+",
        type=Path,
        default=[Path("docs/Naver_Land_0503_1129.xlsx"), Path("docs/Naver_Land_0503_1132.xlsx")],
    )
    parser.add_argument("--year", required=True, type=int, help="Capture year for Naver_Land_{MMDD}_{HHMM}.xlsx files.")
    parser.add_argument("--output", default=Path("data/normalized/naver_land_0503.normalized.json"), type=Path)
    args = parser.parse_args()

    payload = normalize(args.input, args.output, args.year)
    summary = payload["summary"]
    print(
        "normalized "
        f"files={summary['sourceFileCount']} "
        f"listings={summary['listingCount']} "
        f"referenceApartments={summary['referenceApartmentCount']} "
        f"warnings={summary['warningCount']} "
        f"output={args.output.as_posix()}"
    )


if __name__ == "__main__":
    main()
