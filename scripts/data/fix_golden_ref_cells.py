#!/usr/bin/env python3
"""Fix known #REF! cells on Golden Sample after xlsx round-trip."""

from __future__ import annotations

import argparse
import shutil
import subprocess
import sys
from pathlib import Path

# Reuse drive helpers from coverage push.
sys.path.insert(0, str(Path(__file__).resolve().parent))
from push_coverage_via_drive_xlsx import (  # noqa: E402
    GOLDEN_SHEET_ID,
    run,
    upload_xlsx_replace,
)

FIXES = {
    # (district, dong, zone_name) -> {col_header: value}
    ("용산구", "청파동", "청파 1구역"): {"기축 아파트 시세(억)": "31"},
}


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sheet-id", default=GOLDEN_SHEET_ID)
    parser.add_argument("--remote", default="gdrive")
    parser.add_argument("--workdir", type=Path, default=Path("data/tmp_fix_ref"))
    args = parser.parse_args()

    from openpyxl import load_workbook

    workdir = args.workdir
    if workdir.exists():
        shutil.rmtree(workdir)
    workdir.mkdir(parents=True)

    run(
        [
            "rclone",
            "backend",
            "copyid",
            f"{args.remote}:",
            args.sheet_id,
            f"{workdir}/",
            "--drive-export-formats",
            "xlsx",
        ]
    )
    src = next(workdir.glob("*.xlsx"))
    out = workdir / "golden_fixed.xlsx"
    wb = load_workbook(src)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    col = {name: idx + 1 for idx, name in enumerate(headers) if name}
    fixed = 0
    for row in range(2, ws.max_row + 1):
        key = (
            str(ws.cell(row, col["행정구"]).value or ""),
            str(ws.cell(row, col["행정동"]).value or ""),
            str(ws.cell(row, col["구역명"]).value or ""),
        )
        if key not in FIXES:
            continue
        for header, value in FIXES[key].items():
            current = ws.cell(row, col[header]).value
            if current == "#REF!" or current is None or str(current).strip() == "":
                ws.cell(row, col[header], value)
                fixed += 1
            elif str(current) == "#REF!":
                ws.cell(row, col[header], value)
                fixed += 1
    # Also sweep any remaining #REF! in 기축 아파트 시세 from prior CSV if present in workdir sibling - skip.
    # Generic: replace literal #REF! with blank is worse; only known fixes.
    for row in range(2, ws.max_row + 1):
        for header in ("기축 아파트 시세(억)", "비교 기축 아파트"):
            cell = ws.cell(row, col[header])
            if cell.value == "#REF!":
                # leave for known fix only; already handled above
                pass

    wb.save(out)
    print(f"fixed_cells={fixed}")
    upload_xlsx_replace(args.sheet_id, out, remote=args.remote)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
