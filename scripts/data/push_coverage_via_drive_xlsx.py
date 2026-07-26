#!/usr/bin/env python3
"""Add coverage column to Golden Sample via Drive export/import (xlsx), not Sheets API.

rclone shared OAuth often cannot call Sheets API; Drive file update with xlsx works.
"""

from __future__ import annotations

import argparse
import configparser
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))
from normalize_golden_samples import derive_coverage, normalize_stage  # noqa: E402

GOLDEN_SHEET_ID = "1YaZjGX53HGNQyAjLp0AIQFGq-j0IWcfPl5WFgUeUfuY"
COVERAGE_HEADER = "coverage"
STAGE_HEADER = "현재 단계"
SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]


def run(cmd: list[str]) -> None:
    print("+", " ".join(cmd), flush=True)
    completed = subprocess.run(cmd, check=False)
    if completed.returncode != 0:
        raise SystemExit(f"command failed ({completed.returncode}): {' '.join(cmd)}")


def rclone_conf() -> configparser.ConfigParser:
    candidates = [
        Path.home() / "AppData" / "Roaming" / "rclone" / "rclone.conf",
        Path.home() / ".config" / "rclone" / "rclone.conf",
    ]
    for path in candidates:
        if path.exists():
            conf = configparser.ConfigParser()
            conf.read(path, encoding="utf-8")
            return conf
    raise FileNotFoundError("rclone.conf not found")


def drive_credentials(remote: str = "gdrive"):
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials

    conf = rclone_conf()
    section = remote if conf.has_section(remote) else None
    if section is None:
        matches = [s for s in conf.sections() if s == remote or s.endswith(remote)]
        if not matches:
            raise SystemExit(f"rclone remote not found: {remote}")
        section = matches[0]

    # Refresh rclone token first so access_token is fresh.
    run(["rclone", "about", f"{remote}:"])

    conf = rclone_conf()
    token = json.loads(conf.get(section, "token"))
    client_id = conf.get(section, "client_id", fallback="") or None
    client_secret = conf.get(section, "client_secret", fallback="") or None

    creds = Credentials(
        token=token.get("access_token"),
        refresh_token=token.get("refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=client_id,
        client_secret=client_secret,
        scopes=SCOPES,
    )
    if client_id and client_secret and creds.refresh_token and not creds.valid:
        creds.refresh(Request())
    if not creds.token:
        raise SystemExit("No usable Drive access token after rclone refresh")
    return creds


def add_coverage_to_workbook(xlsx_path: Path, out_path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    # Main golden tab is usually first sheet.
    ws = wb.worksheets[0]
    headers = [cell.value for cell in ws[1]]
    try:
        stage_idx = headers.index(STAGE_HEADER) + 1  # 1-based
    except ValueError as exc:
        raise SystemExit(f"header missing on {ws.title}: {STAGE_HEADER}") from exc

    if COVERAGE_HEADER in headers:
        coverage_idx = headers.index(COVERAGE_HEADER) + 1
        inserted = False
    else:
        coverage_idx = stage_idx + 1
        ws.insert_cols(coverage_idx)
        ws.cell(1, coverage_idx, COVERAGE_HEADER)
        inserted = True

    core_n = 0
    sub_n = 0
    rows_updated = 0
    for row in range(2, ws.max_row + 1):
        if all(ws.cell(row, col).value in (None, "") for col in range(1, ws.max_column + 1)):
            continue
        stage_raw = ws.cell(row, stage_idx).value
        stage = normalize_stage(str(stage_raw).strip() if stage_raw is not None else None)
        coverage = derive_coverage(stage)
        ws.cell(row, coverage_idx, coverage)
        rows_updated += 1
        if coverage == "CORE":
            core_n += 1
        else:
            sub_n += 1

    wb.save(out_path)
    return {
        "sheet": ws.title,
        "insertedColumn": inserted,
        "rowsUpdated": rows_updated,
        "core": core_n,
        "sub": sub_n,
        "output": str(out_path),
    }


def upload_xlsx_replace(file_id: str, xlsx_path: Path, remote: str = "gdrive") -> None:
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = drive_credentials(remote)
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    # Convert back to native Google Sheet content.
    updated = (
        service.files()
        .update(
            fileId=file_id,
            media_body=media,
            body={"mimeType": "application/vnd.google-apps.spreadsheet"},
            supportsAllDrives=True,
        )
        .execute()
    )
    print(json.dumps({"uploaded": True, "id": updated.get("id"), "name": updated.get("name")}, ensure_ascii=False))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--sheet-id", default=GOLDEN_SHEET_ID)
    parser.add_argument("--remote", default="gdrive")
    parser.add_argument("--workdir", type=Path, default=Path("data/tmp_coverage_sheet"))
    parser.add_argument("--skip-upload", action="store_true")
    args = parser.parse_args()

    workdir = args.workdir
    if workdir.exists():
        shutil.rmtree(workdir)
    workdir.mkdir(parents=True)

    run(
        [
            "rclone",
            "backend",
            "copyid",
            f"{args.remote}:",
            args.sheet_id,
            f"{workdir}/",
            "--drive-export-formats",
            "xlsx",
        ]
    )
    xlsx_files = sorted(workdir.glob("*.xlsx"))
    if not xlsx_files:
        raise SystemExit(f"no xlsx exported into {workdir}")
    src = xlsx_files[0]
    out = workdir / "golden_with_coverage.xlsx"
    stats = add_coverage_to_workbook(src, out)
    print(json.dumps(stats, ensure_ascii=False, indent=2))

    if not args.skip_upload:
        upload_xlsx_replace(args.sheet_id, out, remote=args.remote)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
