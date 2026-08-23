#!/usr/bin/env python3
"""Apply jaegebal district-status report to Golden Sample 본표.

Updates:
  - 현재 단계 (D) from top meta mapped stage when stageDiff=changed
  - coverage (E) re-derived from new stage when stage changes (default on)
  - 진행현황 column (create if missing) — merged text, new items only already merged in report
  - jaegebal_url

Default dry-run. Pass --write to upload via Drive xlsx.
"""

from __future__ import annotations

import argparse
import configparser
import json
import re
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parent))
from normalize_golden_samples import derive_coverage  # noqa: E402

GOLDEN_SHEET_ID = "1YaZjGX53HGNQyAjLp0AIQFGq-j0IWcfPl5WFgUeUfuY"
STAGE_HEADER = "현재 단계"
COVERAGE_HEADER = "coverage"
PROGRESS_HEADER = "진행현황"
URL_HEADER = "jaegebal_url"
CHECKED_HEADER = "jaegebal_checked_at"
SOURCE_DATE_HEADER = "jaegebal_stage_date"
UPDATED_HEADER = "stage_updated_at"
# D~M band (by header name). Updated cells in this band get amber highlight.
HIGHLIGHT_DM_HEADERS = [
    "현재 단계",
    "coverage",
    "jaegebal_url",
    "매매가",
    "최소 실투자금(억)",
    "최대 실투자금(억)",
    "최소 프리미엄",
    "최대 프리미엄",
    "조합원분양가(84)",
    "총 투자금",
]
HIGHLIGHT_FILL = "FFF2CC"  # soft amber
SCOPES = [
    "https://www.googleapis.com/auth/drive",
    "https://www.googleapis.com/auth/spreadsheets",
]


def run(cmd: list[str]) -> None:
    print("+", " ".join(cmd), flush=True)
    completed = subprocess.run(cmd, check=False)
    if completed.returncode != 0:
        raise SystemExit(f"command failed ({completed.returncode}): {' '.join(cmd)}")


def rclone_conf() -> configparser.ConfigParser:
    for path in (
        Path.home() / "AppData" / "Roaming" / "rclone" / "rclone.conf",
        Path.home() / ".config" / "rclone" / "rclone.conf",
    ):
        if path.exists():
            conf = configparser.ConfigParser()
            conf.read(path, encoding="utf-8")
            return conf
    raise FileNotFoundError("rclone.conf not found")


def drive_credentials(remote: str = "gdrive"):
    from google.oauth2.credentials import Credentials

    run(["rclone", "about", f"{remote}:"])
    conf = rclone_conf()
    section = remote if conf.has_section(remote) else conf.sections()[0]
    token = json.loads(conf.get(section, "token"))
    return Credentials(
        token=token.get("access_token"),
        refresh_token=token.get("refresh_token"),
        token_uri="https://oauth2.googleapis.com/token",
        client_id=conf.get(section, "client_id", fallback="") or None,
        client_secret=conf.get(section, "client_secret", fallback="") or None,
        scopes=SCOPES,
    )


def zone_key(district, dong, zone) -> str:
    return "|".join([str(district or "").strip(), str(dong or "").strip(), str(zone or "").strip()])


def compact_name(value: str) -> str:
    text = re.sub(r"\s+", "", value or "")
    for token in ("(재건축)", "(모아타운)", "재개발", "신통기획", "재건축", "|재개발닷컴", "구역"):
        text = text.replace(token, "")
    return text


def is_confident_match(zone_name: str, page_title: str, badge: str = "") -> bool:
    """Reject obvious search mismatches (금호22→금호21, 신정1→신정1-3, 용산1→남산3)."""
    z = compact_name(zone_name)
    t = compact_name(page_title)
    b = compact_name(badge)
    hay = t + b
    if not z or not t:
        return False
    # compound title like 신정1-3 when zone is 신정1
    if re.search(re.escape(z) + r"-\d", t):
        return False
    if z in hay:
        return True
    zm = re.match(r"^([가-힣]+)(\d+)$", z)
    if zm:
        name, num = zm.group(1), zm.group(2)
        title_nums = re.findall(r"\d+", t)
        if name in hay and num in title_nums and f"{name}{num}-" not in t:
            # require the zone number to appear as the same name+num token roughly
            if f"{name}{num}" in t or (name in t and title_nums.count(num) >= 1 and num in title_nums):
                # reject if title's primary numbered token differs, e.g. 금호21 vs 금호22
                m2 = re.search(re.escape(name) + r"(\d+)", t)
                if m2 and m2.group(1) != num:
                    return False
                return True
    # shared significant tokens (length>=2)
    z_toks = set(re.findall(r"[가-힣]{2,}|\d+", z))
    t_toks = set(re.findall(r"[가-힣]{2,}|\d+", t + b))
    if not z_toks:
        return False
    overlap = z_toks & t_toks
    if z_toks <= t_toks:
        return True
    # if both have numbers, numbers must intersect
    z_nums = {x for x in z_toks if x.isdigit()}
    t_nums = {x for x in t_toks if x.isdigit()}
    if z_nums and t_nums and not (z_nums & t_nums):
        return False
    return len(overlap) >= max(1, len(z_toks) - 1)


def ensure_progress_column(ws) -> int:
    headers = [c.value for c in ws[1]]
    if PROGRESS_HEADER in headers:
        return headers.index(PROGRESS_HEADER) + 1
    # append at end
    col = ws.max_column + 1
    ws.cell(1, col, PROGRESS_HEADER)
    return col


def shade_if_dm(ws, hm: dict[str, int], row: int, header: str, fill: PatternFill, shaded: list) -> None:
    if header not in HIGHLIGHT_DM_HEADERS or header not in hm:
        return
    cell = ws.cell(row, hm[header])
    cell.fill = fill
    shaded.append(f"{header}!{row}")


def apply_to_workbook(
    xlsx: Path,
    out: Path,
    zones: list[dict[str, Any]],
    *,
    day: str,
    update_coverage: bool = True,
    apply_stage_changes: bool = False,
) -> dict[str, Any]:
    wb = load_workbook(xlsx)
    ws = wb.worksheets[0]
    headers = [c.value for c in ws[1]]
    hm = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
    for need in ("행정구", "행정동", "구역명", STAGE_HEADER):
        if need not in hm:
            raise SystemExit(f"missing header {need}")

    # ensure helper cols
    for name in (URL_HEADER, SOURCE_DATE_HEADER, CHECKED_HEADER, UPDATED_HEADER):
        if name not in hm:
            col = ws.max_column + 1
            ws.cell(1, col, name)
            hm[name] = col
    if update_coverage and COVERAGE_HEADER not in hm:
        # place coverage after 현재 단계 when missing
        col = ws.max_column + 1
        ws.cell(1, col, COVERAGE_HEADER)
        hm[COVERAGE_HEADER] = col
    progress_col = ensure_progress_column(ws)
    headers = [c.value for c in ws[1]]
    hm = {str(h).strip(): i + 1 for i, h in enumerate(headers) if h}
    fill = PatternFill(start_color=HIGHLIGHT_FILL, end_color=HIGHLIGHT_FILL, fill_type="solid")

    by_key = {
        z["zoneNaturalKey"]: z
        for z in zones
        if z.get("zoneNaturalKey") and z.get("stageDiff") != "error"
    }
    # also index by zone name only for ad-hoc tests missing district/dong
    by_name: dict[str, dict[str, Any]] = {}
    for z in by_key.values():
        name = (z.get("zoneName") or "").strip()
        if name:
            by_name[name] = z
            by_name[name.replace(" ", "")] = z

    stats = {
        "stageUpdated": 0,
        "coverageUpdated": 0,
        "progressUpdated": 0,
        "urlUpdated": 0,
        "matchedRows": 0,
        "shadedCells": 0,
        "skippedLowConfidence": [],
        "unmatchedReportZones": [],
        "stagePendingApproval": [],
        "skippedMoaStageOnNonMoa": [],
        "promotedToCore": [],
    }

    matched_keys: set[str] = set()
    for r in range(2, ws.max_row + 1):
        zone = ws.cell(r, hm["구역명"]).value
        if not zone:
            continue
        key = zone_key(ws.cell(r, hm["행정구"]).value, ws.cell(r, hm["행정동"]).value, zone)
        z = by_key.get(key) or by_name.get(str(zone).strip()) or by_name.get(str(zone).replace(" ", ""))
        if not z:
            continue
        matched_keys.add(z["zoneNaturalKey"])
        stats["matchedRows"] += 1

        trusted = bool(z.get("forceTrusted")) or is_confident_match(
            str(zone), str(z.get("pageTitle") or ""), str(z.get("badgeText") or "")
        )
        if not trusted:
            stats["skippedLowConfidence"].append(
                {
                    "key": z.get("zoneNaturalKey"),
                    "title": z.get("pageTitle"),
                    "url": z.get("jaegebalUrl"),
                }
            )
            continue

        shaded: list[str] = []

        if z.get("jaegebalUrl"):
            prev = ws.cell(r, hm[URL_HEADER]).value
            if str(prev or "").strip() != str(z["jaegebalUrl"]).strip():
                ws.cell(r, hm[URL_HEADER], z["jaegebalUrl"])
                shade_if_dm(ws, hm, r, URL_HEADER, fill, shaded)
                stats["urlUpdated"] += 1
            else:
                ws.cell(r, hm[URL_HEADER], z["jaegebalUrl"])
        if z.get("topStageDate"):
            ws.cell(r, hm[SOURCE_DATE_HEADER], str(z["topStageDate"]).replace(".", "-"))
        ws.cell(r, hm[CHECKED_HEADER], day)

        stage_written = None
        cur_stage = str(ws.cell(r, hm[STAGE_HEADER]).value or "").strip()
        new_stage = str(z.get("seedfitStageCandidate") or "").strip()
        zone_name = str(zone)
        would_change_stage = bool(new_stage) and (
            (z.get("stageDiff") == "changed")
            or (bool(z.get("forceTrusted")) and cur_stage != new_stage)
        )
        if would_change_stage and "모아" in new_stage and "모아" not in zone_name:
            stats["skippedMoaStageOnNonMoa"].append(
                {"key": z.get("zoneNaturalKey"), "from": cur_stage, "to": new_stage}
            )
            would_change_stage = False
        if would_change_stage and not apply_stage_changes:
            stats["stagePendingApproval"].append(
                {
                    "key": z.get("zoneNaturalKey"),
                    "from": cur_stage,
                    "to": new_stage,
                    "top": z.get("topStageRaw"),
                }
            )
        elif would_change_stage:
            stage_written = new_stage
            ws.cell(r, hm[STAGE_HEADER], stage_written)
            shade_if_dm(ws, hm, r, STAGE_HEADER, fill, shaded)
            if UPDATED_HEADER in hm:
                ws.cell(r, hm[UPDATED_HEADER], day)
            stats["stageUpdated"] += 1

        if update_coverage and COVERAGE_HEADER in hm and stage_written:
            new_cov = derive_coverage(stage_written)
            prev_cov = str(ws.cell(r, hm[COVERAGE_HEADER]).value or "").strip().upper()
            if prev_cov != new_cov:
                ws.cell(r, hm[COVERAGE_HEADER], new_cov)
                shade_if_dm(ws, hm, r, COVERAGE_HEADER, fill, shaded)
                stats["coverageUpdated"] += 1
                if prev_cov == "SUB" and new_cov == "CORE":
                    stats["promotedToCore"].append(z.get("zoneNaturalKey"))

        if z.get("progressMerged"):
            prev = ws.cell(r, progress_col).value
            if str(prev or "") != z["progressMerged"]:
                ws.cell(r, progress_col, z["progressMerged"])
                stats["progressUpdated"] += 1

        stats["shadedCells"] += len(shaded)

    stats["unmatchedReportZones"] = [k for k in by_key if k not in matched_keys]
    wb.save(out)
    stats["progressColumn"] = progress_col
    stats["output"] = str(out)
    return stats


def upload_xlsx_replace(file_id: str, xlsx_path: Path, remote: str = "gdrive") -> None:
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload

    creds = drive_credentials(remote)
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    media = MediaFileUpload(
        str(xlsx_path),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        resumable=True,
    )
    updated = (
        service.files()
        .update(
            fileId=file_id,
            media_body=media,
            body={"mimeType": "application/vnd.google-apps.spreadsheet"},
            supportsAllDrives=True,
        )
        .execute()
    )
    print(json.dumps({"uploaded": True, "id": updated.get("id"), "name": updated.get("name")}, ensure_ascii=False))


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--diff", type=Path, required=True, help="jaegebal_district_status_*.json")
    parser.add_argument("--sheet-id", default=GOLDEN_SHEET_ID)
    parser.add_argument("--remote", default="gdrive")
    parser.add_argument("--workdir", type=Path, default=Path("data/tmp_apply_district_status"))
    parser.add_argument("--write", action="store_true")
    parser.add_argument(
        "--apply-stage-changes",
        action="store_true",
        help="Write 현재 단계/coverage. Default: 진행현황·url만. 단계 변경은 사용자 승인 후 이 플래그로.",
    )
    parser.add_argument(
        "--no-update-coverage",
        action="store_true",
        help="Do not re-derive coverage from stage (default: update coverage).",
    )
    args = parser.parse_args()

    payload = json.loads(args.diff.read_text(encoding="utf-8"))
    zones = list(payload.get("zones") or [])
    plan = {
        "diff": args.diff.as_posix(),
        "zones": len(zones),
        "updateCoverage": not args.no_update_coverage,
        "stageChanges": [
            {"key": z.get("zoneNaturalKey"), "from": z.get("sheetStage"), "to": z.get("seedfitStageCandidate"), "top": z.get("topStageRaw")}
            for z in zones
            if z.get("stageDiff") == "changed"
        ],
        "progressUpdates": [
            {"key": z.get("zoneNaturalKey"), "new": z.get("progressNewLabels"), "merged": z.get("progressMerged")}
            for z in zones
            if z.get("progressMerged")
        ],
        "write": bool(args.write),
        "applyStageChanges": bool(args.apply_stage_changes),
        "coverageScope": "CORE+SUB (input CSV 전 행. coverage로 제외하지 않음)",
    }
    print(json.dumps(plan, ensure_ascii=False, indent=2))
    if plan["stageChanges"] and not args.apply_stage_changes:
        print(
            "현재 단계 변경 "
            f"{len(plan['stageChanges'])}건은 사용자 승인 전 시트에 안 씁니다. "
            "목록을 보여 승인을 받은 뒤 --write --apply-stage-changes 로 재실행하세요. "
            "진행현황·url은 --write 만으로 반영됩니다."
        )
    if not args.write:
        print("dry-run only; re-run with --write after review")
        return 0

    from datetime import datetime
    from zoneinfo import ZoneInfo

    day = datetime.now(ZoneInfo("Asia/Seoul")).strftime("%Y-%m-%d")
    workdir = args.workdir
    if workdir.exists():
        shutil.rmtree(workdir)
    workdir.mkdir(parents=True)
    run(
        [
            "rclone",
            "backend",
            "copyid",
            f"{args.remote}:",
            args.sheet_id,
            f"{workdir}/",
            "--drive-export-formats",
            "xlsx",
        ]
    )
    files = sorted(workdir.glob("*.xlsx"))
    if not files:
        raise SystemExit("no xlsx exported")
    out = workdir / "applied.xlsx"
    stats = apply_to_workbook(
        files[0],
        out,
        zones,
        day=day,
        update_coverage=not args.no_update_coverage,
        apply_stage_changes=bool(args.apply_stage_changes),
    )
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    upload_xlsx_replace(args.sheet_id, out, remote=args.remote)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
